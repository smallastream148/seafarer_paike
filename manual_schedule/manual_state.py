import pandas as pd
from typing import List, Dict, Optional

# 兼容作为包(import manual_schedule.*) 或直接脚本所在目录运行
try:  # 包形式
    from .manual_core import ManualScheduler, TimetableData, PlacedBlock
    from .manual_soft import evaluate_soft
except ImportError:  # 脚本直接执行形式
    from manual_core import ManualScheduler, TimetableData, PlacedBlock  # type: ignore
    from manual_soft import evaluate_soft  # type: ignore

class ManualSession:
    def __init__(self, data: Optional[TimetableData]=None):
        self.data = data or TimetableData()
        self.scheduler = ManualScheduler(self.data)

    def add_block(self, class_id: str, course: str, teacher1: str, teacher2: str|None, date, period:int):
        blk = PlacedBlock(class_id, course, teacher1, teacher2, date, period)
        ok, errs = self.scheduler.add_block(blk)
        return ok, errs

    def undo(self):
        return self.scheduler.remove_last()

    def delete_block(self, idx: int):
        return self.scheduler.delete_block(idx)

    def soft_report(self):
        adjust, details = evaluate_soft(self.scheduler.placed, self.data)
        return adjust, details

    def export_excel(self, path: str, class_id: str|None=None):
        """导出：
        - 若指定 class_id: 仅该班，并生成 1) 明细 2) 班级课表(透视) 3) 教师课时(该班) 4) 课程进度(该班) 5) 软约束(该班)
        - 若不指定: 所有班级；除全局统计外，为每个班生成一个独立 sheet (透视表) 名称: 班级ID
        透视表结构: 行=日期, 列=上午/下午; 单元格 "课程\n教师1[/教师2]"，若多块同一时段则以分号分隔
        """
        all_rows = self.scheduler.export_rows()
        df_all = pd.DataFrame(all_rows)
        if class_id:
            df = df_all[df_all['班级ID']==class_id].copy()
        else:
            df = df_all.copy()
        # 统计教师课时
        teacher_hours = {}
        for _, r in df.iterrows():
            for t in [r['教师1'], r['教师2']]:
                if isinstance(t,str) and t:
                    teacher_hours[t] = teacher_hours.get(t,0)+1
        th_df = pd.DataFrame([
            {'教师':k,'已排课时':v} for k,v in sorted(teacher_hours.items(), key=lambda x:(-x[1],x[0]))
        ])
        # 课程进度
        progress = {}
        for blk in self.scheduler.placed:
            if (not class_id) or blk.class_id==class_id:
                progress.setdefault((blk.class_id, blk.course),0)
                progress[(blk.class_id, blk.course)] += 1
        prog_rows = []
        target_classes = {class_id: self.data.classes[class_id]} if class_id else self.data.classes
        for cid, info in target_classes.items():
            for c in info.courses:
                need = self.data.courses[c].blocks
                got = progress.get((cid,c),0)
                prog_rows.append({'班级ID':cid,'课程':c,'需求块数':need,'已排块数':got,'完成率%': round(got/need*100,2) if need else 0})
        prog_df = pd.DataFrame(prog_rows)
        # 软约束
        if class_id:
            filtered_blocks = [b for b in self.scheduler.placed if b.class_id==class_id]
            adjust, details = evaluate_soft(filtered_blocks, self.data)
        else:
            adjust, details = evaluate_soft(self.scheduler.placed, self.data)
        soft_df = pd.DataFrame([{'soft_total':adjust, **details}])

        # 构造透视表函数
        import datetime as _dt
        def build_pivot(df_class: pd.DataFrame, cid: str):
            if df_class.empty:
                return pd.DataFrame(columns=['日期','上午','下午'])
            # 课程显示: 课程 + 换行 + 教师 /第二教师
            def fmt_row(r):
                if r['教师2']:
                    return f"{r['课程']}\n{r['教师1']}/{r['教师2']}"
                return f"{r['课程']}\n{r['教师1']}"
            dfc = df_class.copy()
            dfc['显示'] = dfc.apply(fmt_row, axis=1)
            # 节次/时段统一映射
            if '节次' in dfc.columns:
                dfc['__period'] = dfc['节次'].map({0:'上午',1:'下午'})
            else:
                dfc['__period'] = dfc['时段']
            # 合并同一日期同一时段多块
            agg = dfc.groupby(['日期','__period'])['显示'].apply(lambda x: '; '.join(x)).unstack('__period')
            # 保证列顺序
            for col in ['上午','下午']:
                if col not in agg.columns:
                    agg[col] = ''
            agg = agg[['上午','下午']].reset_index()
            agg = agg.sort_values('日期')
            return agg

        with pd.ExcelWriter(path, engine='openpyxl') as w:
            # 写基础明细
            df.to_excel(w, sheet_name='排课明细', index=False)
            th_df.to_excel(w, sheet_name='教师课时', index=False)
            prog_df.to_excel(w, sheet_name='课程进度', index=False)
            soft_df.to_excel(w, sheet_name='软约束统计', index=False)
            if class_id:
                pv = build_pivot(df, class_id)
                pv.to_excel(w, sheet_name=f'{class_id}_课表', index=False)
            else:
                for cid in sorted(target_classes.keys()):
                    pv = build_pivot(df[df['班级ID']==cid], cid)
                    pv.to_excel(w, sheet_name=f'{cid}', index=False)
            # 样式美化
            wb = w.book
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            header_font = Font(bold=True, color='223344')
            header_fill = PatternFill('solid', fgColor='dde6f0')
            thin = Side(style='thin', color='c0c7ce')
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            def beautify(ws, wrap=False, freeze=True):
                if freeze:
                    ws.freeze_panes = 'B2'
                max_col = ws.max_column
                # 头样式
                for c in range(1, max_col+1):
                    cell = ws.cell(row=1, column=c)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border
                # 内容样式与列宽
                col_widths = [0]*(max_col+1)
                for r in range(2, ws.max_row+1):
                    for c in range(1, max_col+1):
                        cell = ws.cell(row=r, column=c)
                        if wrap:
                            cell.alignment = Alignment(wrap_text=True, vertical='top')
                        cell.border = border
                        v = cell.value
                        l = 0 if v is None else len(str(v))
                        if l>col_widths[c]:
                            col_widths[c] = min(l, 40)
                for c in range(1, max_col+1):
                    ws.column_dimensions[get_column_letter(c)].width = max(8, col_widths[c]+2)
            # 应用到各 sheet
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                if ws_name.endswith('课表') or ws_name in [*target_classes.keys()]:
                    beautify(ws, wrap=True)
                else:
                    beautify(ws, wrap=False)
        return path

    def export_excel_bytes(self, class_id: str|None=None):
        """与 export_excel 相同逻辑但返回 bytes, 不落地文件."""
        import io
        bio = io.BytesIO()
        # 复用 export_excel 主要逻辑：临时写入内存 Workbook
        all_rows = self.scheduler.export_rows()
        import pandas as pd
        df_all = pd.DataFrame(all_rows)
        if class_id:
            df = df_all[df_all['班级ID']==class_id].copy()
        else:
            df = df_all.copy()
        # 统计教师课时/进度与 soft — 直接调用已有函数部分（复制自 export_excel，保持同步）
        teacher_hours = {}
        for _, r in df.iterrows():
            for t in [r['教师1'], r['教师2']]:
                if isinstance(t,str) and t:
                    teacher_hours[t] = teacher_hours.get(t,0)+1
        th_df = pd.DataFrame([
            {'教师':k,'已排课时':v} for k,v in sorted(teacher_hours.items(), key=lambda x:(-x[1],x[0]))
        ])
        progress = {}
        for blk in self.scheduler.placed:
            if (not class_id) or blk.class_id==class_id:
                progress.setdefault((blk.class_id, blk.course),0)
                progress[(blk.class_id, blk.course)] += 1
        prog_rows = []
        target_classes = {class_id: self.data.classes[class_id]} if class_id else self.data.classes
        for cid, info in target_classes.items():
            for c in info.courses:
                need = self.data.courses[c].blocks
                got = progress.get((cid,c),0)
                prog_rows.append({'班级ID':cid,'课程':c,'需求块数':need,'已排块数':got,'完成率%': round(got/need*100,2) if need else 0})
        prog_df = pd.DataFrame(prog_rows)
        if class_id:
            from manual_schedule.manual_soft import evaluate_soft as _eval_soft
            filtered_blocks = [b for b in self.scheduler.placed if b.class_id==class_id]
            adjust, details = _eval_soft(filtered_blocks, self.data)
        else:
            from manual_schedule.manual_soft import evaluate_soft as _eval_soft
            adjust, details = _eval_soft(self.scheduler.placed, self.data)
        soft_df = pd.DataFrame([{'soft_total':adjust, **details}])

        def build_pivot(df_class: pd.DataFrame, cid: str):
            if df_class.empty:
                return pd.DataFrame(columns=['日期','上午','下午'])
            def fmt_row(r):
                if r['教师2']:
                    return f"{r['课程']}\n{r['教师1']}/{r['教师2']}"
                return f"{r['课程']}\n{r['教师1']}"
            dfc = df_class.copy()
            dfc['显示'] = dfc.apply(fmt_row, axis=1)
            if '节次' in dfc.columns:
                dfc['__period'] = dfc['节次'].map({0:'上午',1:'下午'})
            else:
                dfc['__period'] = dfc['时段']
            agg = dfc.groupby(['日期','__period'])['显示'].apply(lambda x: '; '.join(x)).unstack('__period')
            for col in ['上午','下午']:
                if col not in agg.columns:
                    agg[col] = ''
            agg = agg[['上午','下午']].reset_index().sort_values('日期')
            return agg
        with pd.ExcelWriter(bio, engine='openpyxl') as w:
            df.to_excel(w, sheet_name='排课明细', index=False)
            th_df.to_excel(w, sheet_name='教师课时', index=False)
            prog_df.to_excel(w, sheet_name='课程进度', index=False)
            soft_df.to_excel(w, sheet_name='软约束统计', index=False)
            if class_id:
                build_pivot(df, class_id).to_excel(w, sheet_name=f'{class_id}_课表', index=False)
            else:
                for cid in sorted(target_classes.keys()):
                    build_pivot(df[df['班级ID']==cid], cid).to_excel(w, sheet_name=f'{cid}', index=False)
        bio.seek(0)
        return bio.read()

    def import_from_excel(self, path: str):
        """从自动排课结果 Excel (sheet='排课明细') 导入，填充到当前 session。
        期望列: 班级ID, 课程, 教师1, 教师2, 日期, 节次
        将清空当前已排后再导入; 不进行软硬约束重新验证(假设来源可信)。
        返回: 导入条数
        """
        df = pd.read_excel(path, sheet_name='排课明细')
        # 兼容早期导出列名使用 “时段” (上午/下午) 而非数字节次
        cols = set(df.columns)
        has_period_index = '节次' in cols
        has_period_label = '时段' in cols
        required_base = {'班级ID','课程','教师1','教师2','日期'}
        if not required_base.issubset(cols) or not (has_period_index or has_period_label):
            missing = (required_base - cols)
            if not (has_period_index or has_period_label):
                missing.add('节次/或时段')
            raise ValueError(f'缺少列: {missing}')
        # 若只有时段列则映射到 0/1
        if has_period_label and not has_period_index:
            def _map_period(v):
                if str(v).strip() in ('上午','AM','0'): return 0
                return 1  # 其余视为下午
            df['节次'] = df['时段'].apply(_map_period)
        # 清空
        self.scheduler.placed.clear()
        for _, row in df.iterrows():
            blk = PlacedBlock(
                str(row['班级ID']),
                str(row['课程']),
                str(row['教师1']) if pd.notna(row['教师1']) else '',
                str(row['教师2']) if pd.notna(row['教师2']) and row['教师2']!='' else None,
                pd.to_datetime(row['日期']).date(),
                int(row['节次'])
            )
            # 直接追加, 不重复硬校验(假设自动排课已处理) — 若需严格可改用 add_block
            self.scheduler.placed.append(blk)
        return len(self.scheduler.placed)
